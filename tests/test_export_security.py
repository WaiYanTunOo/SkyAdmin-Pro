"""Excel export must not leak credentials or vault fields."""

import pytest

from skyadmin_pro.database import Database
from skyadmin_pro.services.export import (
    _ALL_EXPORT_COLUMN_MAPS,
    FORBIDDEN_EXPORT_COLUMNS,
    export_to_excel,
)


@pytest.fixture
def db(tmp_path):
    return Database(tmp_path / "export_security.db")


def test_export_column_maps_exclude_forbidden_fields():
    exported_keys = {key for mapping in _ALL_EXPORT_COLUMN_MAPS for key in mapping}
    leaked = exported_keys & FORBIDDEN_EXPORT_COLUMNS
    assert not leaked, f"Forbidden columns in export maps: {sorted(leaked)}"


def test_export_excludes_credential_and_ird_secrets(db, tmp_path, fake_app_dir, monkeypatch):
    monkeypatch.setattr(
        "skyadmin_pro.services.secret_fields.get_machine_id",
        lambda: "TESTMACHINE00001",
    )
    client_id = db.get_or_create_client("Secret Co")
    db.update_client_fields(client_id, ird_password="plaintext-leak-ird")
    db.add_client_credential(
        client_id=client_id,
        credential_type="RD",
        password="rd-secret-value",
        login_id="rd-login-99",
    )

    dest = tmp_path / "export.xlsx"
    export_to_excel(db, dest)
    assert dest.exists()

    import openpyxl

    wb = openpyxl.load_workbook(dest)
    forbidden_headers = {
        "ird password",
        "ird_password",
        "secret_value",
        "password",
        "login id",
        "registration number",
    }
    secret_values = {"plaintext-leak-ird", "rd-secret-value", "rd-login-99"}

    for sheet in wb.worksheets:
        if sheet.max_row < 1:
            continue
        headers = [str(cell.value or "").strip().lower() for cell in sheet[1]]
        assert not forbidden_headers.intersection(headers), f"Forbidden header in sheet {sheet.title!r}: {headers}"
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for value in row:
                text = str(value or "").strip()
                if not text:
                    continue
                assert text not in secret_values, f"Secret value leaked in sheet {sheet.title!r}: {text!r}"


def test_monthly_report_export_excludes_secrets(db, tmp_path, fake_app_dir, monkeypatch):
    monkeypatch.setattr(
        "skyadmin_pro.services.secret_fields.get_machine_id",
        lambda: "TESTMACHINE00001",
    )
    client_id = db.get_or_create_client("Report Co")
    db.update_client_fields(client_id, ird_password="monthly-report-leak")
    with db.connection() as conn:
        conn.execute(
            "INSERT INTO documents (client_id, document_type, start_date, amount) VALUES (?, ?, ?, ?)",
            (client_id, "Visa", "2026-08-01", "5000"),
        )

    dest = tmp_path / "monthly.xlsx"
    from skyadmin_pro.services.export import export_monthly_report

    export_monthly_report(db, 2026, 8, dest)
    assert dest.exists()

    import openpyxl

    wb = openpyxl.load_workbook(dest)
    sheet = wb["Pipeline"]
    flat = {str(value) for row in sheet.iter_rows(values_only=True) for value in row if value is not None}
    assert "monthly-report-leak" not in flat
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    assert headers == ["No.", "Date", "Client", "Service", "Amount"]
    assert "ird_password" not in {h.lower() for h in headers}
    assert "password" not in {h.lower() for h in headers}
